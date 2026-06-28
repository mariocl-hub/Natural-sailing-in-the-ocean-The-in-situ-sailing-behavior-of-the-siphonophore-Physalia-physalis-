from ultralytics import YOLO
import cv2
from tqdm import tqdm
import csv
from collections import defaultdict
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

# Diccionario con todas las configuraciones de videos
VIDEO_CONFIGS = {
    # ===== VIDEO 1 - 30_07_2025_10AM =====
    "video1_clip1": {
        "input": r"C:\Users\Usuario\Documents\Mario\Estudios\Master Universitario\TFM\database\Video 1 (30_07_2025_10AM)\DJI_20250730102351_0001_D\DJI_20250730102351_0001_D_detected_cropped.mp4",
        "output": r"C:\Users\Usuario\Documents\Mario\Estudios\Master Universitario\TFM\database\Video 1 (30_07_2025_10AM)\DJI_20250730102351_0001_D\DJI_20250730102351_0001_D_detected_cropped_classified.mp4",
        "epoch": 807186231
    },
    "video1_clip2": {
        "input": r"C:\Users\Usuario\Documents\Mario\Estudios\Master Universitario\TFM\database\Video 1 (30_07_2025_10AM)\DJI_20250730102351_0002_D\DJI_20250730102740_0002_D_detected_cropped.mp4",
        "output": r"C:\Users\Usuario\Documents\Mario\Estudios\Master Universitario\TFM\database\Video 1 (30_07_2025_10AM)\DJI_20250730102351_0002_D\DJI_20250730102740_0002_D_detected_cropped_classified.mp4",
        "epoch": 807186460
    },
    "video1_clip3": {
        "input": r"C:\Users\Usuario\Documents\Mario\Estudios\Master Universitario\TFM\database\Video 1 (30_07_2025_10AM)\DJI_20250730102351_0003_D\DJI_20250730103129_0003_D_detected_cropped.mp4",
        "output": r"C:\Users\Usuario\Documents\Mario\Estudios\Master Universitario\TFM\database\Video 1 (30_07_2025_10AM)\DJI_20250730102351_0003_D\DJI_20250730103129_0003_D_detected_cropped_classified.mp4",
        "epoch": 807186689
    },
    "video1_clip4": {
        "input": r"C:\Users\Usuario\Documents\Mario\Estudios\Master Universitario\TFM\database\Video 1 (30_07_2025_10AM)\DJI_20250730102351_0004_D\DJI_20250730103517_0004_D_detected_cropped.mp4",
        "output": r"C:\Users\Usuario\Documents\Mario\Estudios\Master Universitario\TFM\database\Video 1 (30_07_2025_10AM)\DJI_20250730102351_0004_D\DJI_20250730103517_0004_D_detected_cropped_classified.mp4",
        "epoch": 807186918
    },
    
    # ===== VIDEO 2 - 30_07_2025_11AM =====
    "video2_clip1": {
        "input": r"C:\Users\Usuario\Documents\Mario\Estudios\Master Universitario\TFM\database\Video 2 (30_07_2025_11AM)\DJI_20250730110740_0003_D\DJI_20250730110740_0003_D_detected_cropped.MP4",
        "output": r"C:\Users\Usuario\Documents\Mario\Estudios\Master Universitario\TFM\database\Video 2 (30_07_2025_11AM)\DJI_20250730110740_0003_D\DJI_20250730110740_0003_D_detected_cropped_classified.mp4",
        "epoch": 807188860
    },
    "video2_clip2": {
        "input": r"C:\Users\Usuario\Documents\Mario\Estudios\Master Universitario\TFM\database\Video 2 (30_07_2025_11AM)\DJI_20250730111129_0004_D\DJI_20250730111129_0004_D_detected_cropped.MP4",
        "output": r"C:\Users\Usuario\Documents\Mario\Estudios\Master Universitario\TFM\database\Video 2 (30_07_2025_11AM)\DJI_20250730111129_0004_D\DJI_20250730111129_0004_D_detected_cropped_classified.mp4",
        "epoch": 807189089
    },
    "video2_clip3": {
        "input": r"C:\Users\Usuario\Documents\Mario\Estudios\Master Universitario\TFM\database\Video 2 (30_07_2025_11AM)\DJI_20250730111517_0005_D\DJI_20250730111517_0005_D_detected_cropped.mp4",
        "output": r"C:\Users\Usuario\Documents\Mario\Estudios\Master Universitario\TFM\database\Video 2 (30_07_2025_11AM)\DJI_20250730111517_0005_D\DJI_20250730111517_0005_D_detected_cropped_classified.mp4",
        "epoch": 807189317
    },
    "video2_clip4": {
        "input": r"C:\Users\Usuario\Documents\Mario\Estudios\Master Universitario\TFM\database\Video 2 (30_07_2025_11AM)\DJI_20250730111906_0006_D\DJI_20250730111906_0006_D_detected_cropped.mp4",
        "output": r"C:\Users\Usuario\Documents\Mario\Estudios\Master Universitario\TFM\database\Video 2 (30_07_2025_11AM)\DJI_20250730111906_0006_D\DJI_20250730111906_0006_D_detected_cropped_classified.mp4",
        "epoch": 807189546
    },
    "video2_clip5": {
        "input": r"C:\Users\Usuario\Documents\Mario\Estudios\Master Universitario\TFM\database\Video 2 (30_07_2025_11AM)\DJI_20250730112255_0007_D\DJI_20250730112255_0007_D_detected_cropped.mp4",
        "output": r"C:\Users\Usuario\Documents\Mario\Estudios\Master Universitario\TFM\database\Video 2 (30_07_2025_11AM)\DJI_20250730112255_0007_D\DJI_20250730112255_0007_D_detected_cropped_classified.mp4",
        "epoch": 807189775
    },

    # ===== VIDEO 3 - 30_07_2025_11AM =====
    "video3_clip1": {
        "input": r"C:\Users\Usuario\Documents\Mario\Estudios\Master Universitario\TFM\database\Video 3 (30_07_2025_11_30AM)\DJI_20250730113843_0001_D\DJI_20250730113843_0001_D-002_detected_cropped.MP4",
        "output": r"C:\Users\Usuario\Documents\Mario\Estudios\Master Universitario\TFM\database\Video 3 (30_07_2025_11_30AM)\DJI_20250730113843_0001_D\DJI_20250730113843_0001_D_detected_cropped_classified.mp4",
        "epoch": 807190723
    },
    "video3_clip2": {
        "input": r"C:\Users\Usuario\Documents\Mario\Estudios\Master Universitario\TFM\database\Video 3 (30_07_2025_11_30AM)\DJI_20250730114231_0002_D\DJI_20250730114231_0002_D-001_detected_cropped.MP4",
        "output": r"C:\Users\Usuario\Documents\Mario\Estudios\Master Universitario\TFM\database\Video 3 (30_07_2025_11_30AM)\DJI_20250730114231_0002_D\DJI_20250730114231_0002_D_detected_cropped_classified.mp4",
        "epoch": 807190952
    },
    "video3_clip3": {
        "input": r"C:\Users\Usuario\Documents\Mario\Estudios\Master Universitario\TFM\database\Video 3 (30_07_2025_11_30AM)\DJI_20250730114620_0003_D\DJI_20250730114620_0003_D-003_detected_cropped.MP4",
        "output": r"C:\Users\Usuario\Documents\Mario\Estudios\Master Universitario\TFM\database\Video 3 (30_07_2025_11_30AM)\DJI_20250730114620_0003_D\DJI_20250730114620_0003_D_detected_cropped_classified.mp4",
        "epoch": 807191180
    },
    "video3_clip4": {
        "input": r"C:\Users\Usuario\Documents\Mario\Estudios\Master Universitario\TFM\database\Video 3 (30_07_2025_11_30AM)\DJI_20250730115009_0004_D\DJI_20250730115009_0004_D-004_detected_cropped.MP4",
        "output": r"C:\Users\Usuario\Documents\Mario\Estudios\Master Universitario\TFM\database\Video 3 (30_07_2025_11_30AM)\DJI_20250730115009_0004_D\DJI_20250730115009_0004_D_detected_cropped_classified.mp4",
        "epoch": 807191409
    },
    "video3_clip5": {
        "input": r"C:\Users\Usuario\Documents\Mario\Estudios\Master Universitario\TFM\database\Video 3 (30_07_2025_11_30AM)\DJI_20250730115358_0005_D\DJI_20250730115358_0005_D-005_detected_cropped.MP4",
        "output": r"C:\Users\Usuario\Documents\Mario\Estudios\Master Universitario\TFM\database\Video 3 (30_07_2025_11_30AM)\DJI_20250730115358_0005_D\DJI_20250730115358_0005_D_detected_cropped_classified.mp4",
        "epoch": 807191638
    },
}

clip_names = list(VIDEO_CONFIGS.keys())
print("Available clips:",",".join(clip_names))
#selected_clip = input("Select a clip to process: ".lower().strip())

#Load YOLO model
model = YOLO(r"C:\Users\Usuario\Documents\Mario\Estudios\Master Universitario\TFM\programación\best_model_clasifier.pt")

def get_dominant_class(frames_data):
    """
    Dado un dict {clase: [confianzas]}, devuelve la clase dominante.
    Criterio: mayor frecuencia; en empate, mayor confianza media.
    """
    best_class = None
    best_count = -1
    best_conf = -1.0

    for cls, confs in frames_data.items():
        count = len(confs)
        mean_conf = sum(confs) / count
        if count > best_count or (count == best_count and mean_conf > best_conf):
            best_class = cls
            best_count = count
            best_conf = mean_conf

    return best_class, best_count, best_conf


def process_video(config, model):
    video_path = config["input"]
    output_path = config["output"]
    epoch_start = config["epoch"]

    cap = cv2.VideoCapture(video_path)
    fps = cap.get(cv2.CAP_PROP_FPS)

    # ── Video writer ──────────────────────────────────────────────
    fourcc = cv2.VideoWriter_fourcc(*'mp4v')
    width  = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
    height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
    out = cv2.VideoWriter(output_path, fourcc, fps, (width, height))

    if not out.isOpened():
        print(f"Error opening output: {output_path}")
        return

    print(f"Processing: {video_path}")

    total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))

    # ── Acumulador por segundo ────────────────────────────────────
    # seconds_data[segundo] = {clase: [confianzas]}
    seconds_data = defaultdict(lambda: defaultdict(list))

    # ── Inferencia frame a frame ──────────────────────────────────
    for frame_id in tqdm(range(total_frames), desc="Processing frames"):
        ret, frame = cap.read()
        if not ret:
            break

        results = model(frame)
        r = results[0]

        if r.probs is not None:
            cls       = int(r.probs.top1)
            conf      = float(r.probs.top1conf)
            class_name = model.names[cls]

            second = int(frame_id // fps)          # segundo al que pertenece este frame
            seconds_data[second][class_name].append(conf)

        annotated_frame = r.plot()
        out.write(annotated_frame)

    cap.release()
    out.release()

    # ── Generar Excel ─────────────────────────────────────────────
    xlsx_path = output_path.replace(".mp4", "_per_second.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Per Second Classification"

    # Cabecera
    headers = ["Second", "Timestamp (s)", "Unix Epoch", "Dominant Class",
               "Frame Count", "Mean Confidence", "All Classes (count)"]
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Colores alternos por fila
    fill_even = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    fill_odd  = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    for row_idx, second in enumerate(sorted(seconds_data.keys()), start=2):
        dominant_cls, frame_count, mean_conf = get_dominant_class(seconds_data[second])
        unix_epoch = epoch_start + second

        # Resumen de todas las clases detectadas en ese segundo
        all_classes_summary = " | ".join(
            f"{cls}:{len(confs)}" for cls, confs in sorted(seconds_data[second].items())
        )

        row_data = [
            second,
            second,           # Timestamp en segundos (relativo al clip)
            unix_epoch,
            dominant_cls,
            frame_count,
            round(mean_conf, 4),
            all_classes_summary,
        ]

        fill = fill_even if row_idx % 2 == 0 else fill_odd
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")

    # Ajustar ancho de columnas automáticamente
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    wb.save(xlsx_path)
    print(f"Excel guardado: {xlsx_path}")
    print(f"Finished: {output_path}")


for name, config in VIDEO_CONFIGS.items():
    print(f"\n--- {name} ---")
    process_video(config, model)